# -*- coding: utf-8 -*-
import requests, time, os
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
import openpyxl as excel
import datetime
import wx
import wx.xrc

'''
Company对象为公司对象，可获取系统内某一单的所有信息，新建时需输入一个单号
还有一个很大的问题，是用户界面需要加上最终审核建
一键审核功能也需要加上
但是一键上传功能确实比较鸡肋，倾向于不要添加
返金，中介费需要确切金额，而且和AD并不是同时得到，不能加



        这里是吐槽部分

            首先需要明确现在公司还有很多制度其实是不完善的。虽无伤大雅，但是对程序完美编写造成了很多问题。此处记录一下

                1.理论上兼职是不能上传返金的，但是和机构合作的是可以上传的
                2.返金人没有的是不能返送的   此功能还未实现




'''


class Company(object):
    # 此处规定需传入参数，只有number
    def __init__(self, number):
        self.number = number
        print(number)

        # 返回顾客id 物件名 管理公司 状态
        client_statues = self.get_customer_id()

        self.customer_id = client_statues[0]
        self.client_statues = client_statues[3]

        if client_statues[3] == '该订单未上传或读取问题':

            # 申请金状态
            self.applay_fee_statues = '该订单未上传或读取问题'
            # 申请金金额
            self.applay_fee_money = '该订单未上传或读取问题'
            # 申请金备注
            self.applay_fee_ps = '该订单未上传或读取问题'
            # 返金金额
            self.fan_jin = '该订单未上传或读取问题'

            # AD金额状态
            self.AD_statue = '该订单未上传或读取问题'

            self.money_id = '该订单未上传或读取问题'
            self.ke_shi = '该订单未上传或读取问题'
            self.role = '该订单未上传或读取问题'
            self.check_statue = '该订单未上传或读取问题'
            self.values = '该订单未上传或读取问题'

            self.find_number = '该订单未上传或读取问题'
            self.statute = '该订单未上传或读取问题'
            self.date = '该订单未上传或读取问题'
            self.dandang = '该订单未上传或读取问题'
            self.wu_jian_ming = '该订单未上传或读取问题'
            self.client = '该订单未上传或读取问题'
            self.tel = '该订单未上传或读取问题'
            self.company = '该订单未上传或读取问题'
            self.cai_wu_statues = '该订单未上传或读取问题'
        else:
            applay_fee_list = self.get_client_system_values()
            # 申请金状态
            self.applay_fee_statues = applay_fee_list[0]
            # 申请金金额
            self.applay_fee_money = applay_fee_list[1]
            # 申请金备注
            self.applay_fee_ps = applay_fee_list[2]
            # 返金金额
            self.fan_jin = applay_fee_list[3]
            if self.fan_jin == "0":
                self.fan_jin = ""
            # AD金额状态
            self.AD_statue = applay_fee_list[4]
            self.system_agency_fee = applay_fee_list[5]

            money_id_statues = self.get_moey_id()
            self.money_id = money_id_statues[0]
            self.ke_shi = money_id_statues[1]
            self.role = money_id_statues[2]
            self.check_statue = money_id_statues[3]
            self.values = self.get_cai_wu_shu_ju()

            value = self.get_company_value()
            self.find_number = value[0]
            self.statute = value[1]
            self.date = value[2]
            self.dandang = value[3]
            self.wu_jian_ming = value[4]
            self.client = value[5]
            self.tel = value[6]
            self.company = value[8]
            self.cai_wu_statues = self.get_cai_wu_zhuang_tai()

    def get_attr(self):
        attr = {"costomer": self.customer_id, "client_statues": self.client_statues,
                "applay_fee_statues": self.applay_fee_statues, "applay_fee_money": self.applay_fee_money,
                "applay_fee_ps": self.applay_fee_ps, "fan_jin": self.fan_jin, "values": self.values,
                "money_id": self.money_id, "ke_shi": self.ke_shi}

    def login_res(self, name, password):
        login_url = 'http://hy.ayqiandu.net/index/login/index.html'
        formData = {'uname': name, 'passwd': password}
        hea = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) PhantomJS/41.0.2272.118 Safari/537.36'}
        s = requests.Session()
        s.post(login_url, data=formData, headers=hea)
        return s

    def get_moey_id(self):
        s = self.login_res("0188", "0123456789")
        html = s.get('http://hy.ayqiandu.net/index/caiwu/index.html?bumen_id=&user_id=&zt=&xm=&wjfh=' + str(
            self.number) + '&cwshsj=&cwshsjend=').text
        soup = BeautifulSoup(html, features="lxml")

        if "LP" in str(self.number):
            tds = soup.find_all(name="td")

            for i in range(len(tds)):

                if tds[i].get_text().strip() == self.number.strip():
                    money_id = tds[i - 1].get_text()
                    # 调整下行的find_next_siblings()[number]中number可以得到对应数据 0为科室，1为社员或兼职，2为名字，3为客户姓名，4为客户状态
                    ke_shi = tds[i + 1].get_text()
                    role = tds[i + 2].get_text()
                    check_statue = tds[i + 5].get_text()

                    return [money_id, ke_shi, role, check_statue]
            return ['该订单未上传', '', "", "该订单未上传"]
        else:

            if soup.find_all(name='td', text=self.number):
                money_id = soup.find_all(name='td', text=self.number)[0].find_previous_sibling().get_text()
                ke_shi = soup.find_all(name='td', text=self.number)[0].find_next_siblings()[0].get_text()
                role = soup.find_all(name='td', text=self.number)[0].find_next_siblings()[1].get_text()
                check_statue = soup.find_all(name='td', text=self.number)[0].find_next_siblings()[4].get_text()
                return [money_id, ke_shi, role, check_statue]
            else:
                return ["该订单未上传", "", "", ""]

        # 想获取更多客户管理系统就在这里更改吧
        #

    def get_cai_wu_shu_ju(self):
        money_id = self.money_id
        s = self.login_res("0188", "0123456789")

        def get_values(money_id, s):
            value = []
            html = s.get('http://hy.ayqiandu.net/index/caiwu/info/id/' + str(money_id) + '.html').text
            soup = BeautifulSoup(html, features="lxml")
            tbodys = soup.find_all('tbody')
            # 下头这俩大兄弟是入金和出金的容纳表格
            # 值得注意的是使用find寻找到的元素依然可以用find进行继续查找

            tds_in1 = tbodys[1]
            tds_in2 = tbodys[2]

            # 每一行有12个td，整个列表里最后一个为统计的金额，请勿读取

            td1s = tds_in1.find_all('td')
            td2s = tds_in2.find_all('td')

            for i in range(len(td1s)):
                if (i - 1) % 13 == 0:
                    mid = []
                    money = td1s[i].get_text().strip()
                    type = td1s[i + 1].get_text().strip()
                    in_out = td1s[i + 2].get_text().strip()
                    qing_qiu_shu = td1s[i + 3].get_text().strip()
                    qi_yue_shu = td1s[i + 4].get_text().strip()
                    time = td1s[i + 9].get_text().strip()
                    check_person = td1s[i + 10].get_text().strip()
                    id = td1s[i - 1].get_text().strip()
                    # 0为类型（AD 其他 中介费 等等） 1为出入金 2为金额 3为请求书 4为契约书 5为审核时间 6为审核人 7为id
                    mid.append(type)
                    mid.append(in_out)
                    mid.append(money)
                    mid.append(qing_qiu_shu)
                    mid.append(qi_yue_shu)
                    mid.append(time)
                    mid.append(check_person)
                    mid.append(id)
                    value.append(mid)

            for i in range(len(td2s)):
                if (i - 1) % 13 == 0:
                    mid = []
                    money = td2s[i].get_text().strip()
                    type = td2s[i + 1].get_text().strip()
                    in_out = td2s[i + 2].get_text().strip()
                    qing_qiu_shu = td2s[i + 3].get_text().strip()
                    qi_yue_shu = td2s[i + 4].get_text().strip()
                    time = td2s[i + 9].get_text().strip()
                    check_person = td2s[i + 10].get_text().strip()
                    id = td2s[i - 1].get_text().strip()
                    # 0为类型（AD 其他 中介费 等等） 1为出入金 2为金额 3为请求书 4为契约书 5为审核时间 6为审核人7为id
                    mid.append(type)
                    mid.append(in_out)
                    mid.append(money)
                    mid.append(qing_qiu_shu)
                    mid.append(qi_yue_shu)
                    mid.append(time)
                    mid.append(check_person)
                    mid.append(id)
                    value.append(mid)
            return value

        return get_values(money_id, s)

    def get_company_value(self):
        money_id = self.get_moey_id()[0]
        print(money_id)
        s = self.login_res("0188", "0123456789")

        def get_values(money_id, s):
            value = []
            html = s.get('http://hy.ayqiandu.net/index/caiwu/info/id/' + str(money_id) + '.html').text
            soup = BeautifulSoup(html, features="lxml")
            tbodys = soup.find_all('tbody')
            # 下头这俩大兄弟是入金和出金的容纳表格
            # 值得注意的是使用find寻找到的元素依然可以用find进行继续查找
            tds_in = tbodys[0]

            # 每一行有12个td，整个列表里最后一个为统计的金额，请勿读取
            tds = tds_in.find_all('td')

            for i in range(len(tds)):
                value.append(tds[i].get_text().strip())

            return value

        return get_values(money_id, s)

    def AD_upload(self, money, date, driver, name):

        s = self.login_res('账号', '密码')
        money_id = self.money_id

        driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/" + str(money_id) + ".html")
        # 下面两部进入添加界面
        tds = driver.find_elements_by_tag_name("td")
        if tds[0].text == "190107":
            pass
        else:
            cai_wu_sheng_qing = driver.find_element_by_link_text("添加财务申请")
            cai_wu_sheng_qing.click()
            time.sleep(1)
            # 开始添加数据

            # 添加金额
            je = driver.find_element_by_name("je")
            je.send_keys(money)

            # 选择种类
            kind = driver.find_element_by_name('type')
            kind_lei = Select(kind)
            kind_lei.select_by_index(1)

            # 选择类别
            lei_bie = driver.find_element_by_name('qqslx')
            lei_bie_lei = Select(lei_bie)
            lei_bie_lei.select_by_index(1)

            # 上传文件
            upload = driver.find_element_by_name('qqs')

            upload.send_keys('C:\\Users\\user\\Desktop\\upload\\pdf\\' + name + "\\pdf\\总表.pdf")
            # 日期
            date_value = driver.find_element_by_class_name('date').find_element_by_tag_name('input')
            date_value.click()

            date_value.send_keys(str(date))

            date_value.send_keys(Keys.ENTER)

            nei_rong = driver.find_element_by_name('qqnr')
            nei_rong_lei = Select(nei_rong)
            nei_rong_lei.select_by_index(1)

            # 添加
            btn = driver.find_element_by_id("caiwuadd")
            # btn=driver.find_element_by_xpath("/html/body/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/div[@class='zygl'][1]/form[@id='[object HTMLInputElement]']/div[@class='form-group'][2]/div[@class='col-xs-12 ']/button[@id='caiwuadd']")
            # btn=driver.find_elements_by_tag_name('button')

            btn.click()
            time.sleep(5)

    def agency_fee_upload(self, money, date, driver):

        s = self.login_res('账号', '密码')
        money_id = self.money_id

        if money_id == "该订单未上传":
            pass
        else:

            driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/" + str(money_id) + ".html")
            # 下面两部进入添加界面
            cai_wu_sheng_qing = driver.find_element_by_link_text("添加财务申请")
            cai_wu_sheng_qing.click()
            time.sleep(1)
            # 开始添加数据

            # 添加金额
            je = driver.find_element_by_name("je")
            je.send_keys(money)

            # 选择种类
            kind = driver.find_element_by_name('type')
            kind_lei = Select(kind)
            kind_lei.select_by_index(1)

            # 选择类别
            lei_bie = driver.find_element_by_name('qqslx')
            lei_bie_lei = Select(lei_bie)
            lei_bie_lei.select_by_index(3)

            # 上传文件
            # upload=driver.find_element_by_name('qqs')

            # upload.send_keys('C:\\Users\\user\\Desktop\\upload\\agency_fee\\'+str(self.number)+".png")
            # 日期
            date_value = driver.find_element_by_class_name('date').find_element_by_tag_name('input')
            date_value.click()

            date_value.send_keys(str(date))

            date_value.send_keys(Keys.ENTER)

            nei_rong = driver.find_element_by_name('qqnr')
            nei_rong_lei = Select(nei_rong)
            nei_rong_lei.select_by_index(0)

            # 添加
            btn = driver.find_element_by_id("caiwuadd")
            # btn=driver.find_element_by_xpath("/html/body/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/div[@class='zygl'][1]/form[@id='[object HTMLInputElement]']/div[@class='form-group'][2]/div[@class='col-xs-12 ']/button[@id='caiwuadd']")
            # btn=driver.find_elements_by_tag_name('button')

            btn.click()
            time.sleep(5)

    def refund_upload(self, money, date, driver):

        s = self.login_res('账号', '密码')
        money_id = self.money_id

        if money_id == "该订单未上传":
            pass
        elif "LP" in str(self.number):
            pass
        else:
            print(self.role)
            if self.role == "	兼职":
                pass
            else:
                driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/" + str(money_id) + ".html")
                # 下面两部进入添加界面
                cai_wu_sheng_qing = driver.find_element_by_link_text("添加财务申请")
                cai_wu_sheng_qing.click()
                time.sleep(1)
                # 开始添加数据

                # 添加金额
                je = driver.find_element_by_name("je")
                je.send_keys(money)

                # 选择种类
                kind = driver.find_element_by_name('type')
                kind_lei = Select(kind)
                kind_lei.select_by_index(2)

                # 日期
                date_value = driver.find_element_by_class_name('date').find_element_by_tag_name('input')
                date_value.click()

                date_value.send_keys(str(date))

                date_value.send_keys(Keys.ENTER)

                nei_rong = driver.find_element_by_name('qqnr')
                nei_rong_lei = Select(nei_rong)
                nei_rong_lei.select_by_index(0)

                # 添加
                btn = driver.find_element_by_id("caiwuadd")
                # btn=driver.find_element_by_xpath("/html/body/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/div[@class='zygl'][1]/form[@id='[object HTMLInputElement]']/div[@class='form-group'][2]/div[@class='col-xs-12 ']/button[@id='caiwuadd']")
                # btn=driver.find_elements_by_tag_name('button')

                btn.click()
                time.sleep(5)

    def affair_upload(self, date, driver):

        s = self.login_res('账号', '密码')
        money_id = self.money_id

        if money_id == "该订单未上传":
            pass
        else:

            driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/" + str(money_id) + ".html")
            # 下面两部进入添加界面
            cai_wu_sheng_qing = driver.find_element_by_link_text("添加财务申请")
            cai_wu_sheng_qing.click()
            time.sleep(1)
            # 开始添加数据

            # 添加金额
            je = driver.find_element_by_name("je")
            je.send_keys("10000")

            # 选择种类
            kind = driver.find_element_by_name('type')
            kind_lei = Select(kind)
            kind_lei.select_by_index(2)

            # 日期
            date_value = driver.find_element_by_class_name('date').find_element_by_tag_name('input')
            date_value.click()

            date_value.send_keys(str(date))

            date_value.send_keys(Keys.ENTER)

            nei_rong = driver.find_element_by_name('qqnr')
            nei_rong_lei = Select(nei_rong)
            nei_rong_lei.select_by_index(2)

            # 添加
            btn = driver.find_element_by_id("caiwuadd")
            # btn=driver.find_element_by_xpath("/html/body/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/div[@class='zygl'][1]/form[@id='[object HTMLInputElement]']/div[@class='form-group'][2]/div[@class='col-xs-12 ']/button[@id='caiwuadd']")
            # btn=driver.find_elements_by_tag_name('button')

            btn.click()
            time.sleep(5)

    def get_cai_wu_zhuang_tai(self):

        zhongjiefei = []
        xian_jin_AD = []
        AD = []
        in_qi_ta = []
        fan_jin = []
        out_qi_ta = []
        if "LP" in str(self.number):
            mid_number = str(self.number)[2:6]
            print(mid_number)
        else:
            mid_number = str(self.number)[0:4]
        if int(mid_number) >= 1907:
            k = 0.3
        else:
            k = 0.5
        for i in range(len(self.values)):

            if self.values[i][1] == "入金":
                if self.values[i][0] == "中介费":
                    zhongjiefei.append(self.values[i])

                elif self.values[i][0] == "现金AD":
                    xian_jin_AD.append(self.values[i])

                elif self.values[i][0] == "AD":
                    AD.append(self.values[i])

                elif self.values[i][0] == "其他":
                    in_qi_ta.append(self.values[i])

            elif self.values[i][1] == "出金":
                if self.values[i][0] == "返金":
                    fan_jin.append(self.values[i])

                elif self.values[i][0] == "其他":
                    out_qi_ta.append(self.values[i])

        # mid_value用于存放中间值
        mid_value = {}

        # 判断中介费状态
        if len(zhongjiefei) == 0:
            if self.system_agency_fee == "0" or "":
                zhong_jie_fei_statues = "无中介费"
            else:
                zhong_jie_fei_statues = '中介费未上传'
            mid_value["中介费"] = ['', zhong_jie_fei_statues]
        elif len(zhongjiefei) == 1:
            zhong_jie_fei_statues = "仅有一笔中介费"
            mid_value["中介费"] = [zhongjiefei[0][2], zhong_jie_fei_statues]
        else:
            zhong_jie_fei_statues = "有" + str(len(zhongjiefei)) + "笔中介费"
            mid_value["中介费"] = ['', zhong_jie_fei_statues]

        # 判断现金AD状态
        if len(xian_jin_AD) == 0:
            xian_jin_AD_statues = '现金AD未上传'
            mid_value["现金AD"] = ['', xian_jin_AD_statues]
        elif len(xian_jin_AD) == 1:
            xian_jin_AD_statues = "仅有一笔现金AD"
            mid_value["现金AD"] = [xian_jin_AD[0][2], xian_jin_AD_statues]
        else:
            xian_jin_AD_statues = "有" + str(len(xian_jin_AD)) + "笔现金AD"
            mid_value["现金AD"] = ['', xian_jin_AD_statues]

        # 判断AD状态
        if len(AD) == 0:
            AD_statues = 'AD未上传'
            if len(self.AD_statue) == 1:
                AD_statues = "AD为0"
            mid_value["AD"] = ['', AD_statues]
        elif len(AD) == 1:
            AD_statues = "仅有一笔AD"
            mid_value["AD"] = [AD[0][2], AD_statues]
        else:
            AD_statues = "有" + str(len(AD)) + "笔AD"
            mid_value["AD"] = ['', AD_statues]

        # 判断返金状态
        if len(fan_jin) == 0:
            if self.fan_jin == "" or self.fan_jin == 0:
                fan_jin_statues = '此单无返金'
            else:
                fan_jin_statues = '返金未上传'
            mid_value["返金"] = ['', fan_jin_statues]

        elif len(fan_jin) == 1:
            fan_jin_statues = "仅有一笔返金"
            mid_value["返金"] = [fan_jin[0][2], fan_jin_statues]

        else:
            fan_jin_statues = "有" + str(len(fan_jin)) + "笔返金"
            mid_value["返金"] = ['', fan_jin_statues]

        # 判断总务扣款状态
        if len(out_qi_ta) == 0:
            out_qi_ta_statues = '总务扣款未上传'
            mid_value["总务扣款"] = ['', out_qi_ta_statues]
        elif len(out_qi_ta) == 1:
            out_qi_ta_statues = "仅有一笔总务扣款"
            mid_value["总务扣款"] = [out_qi_ta[0][2], out_qi_ta_statues]
        else:
            out_qi_ta_statues = "有" + str(len(out_qi_ta)) + "笔总务扣款"
            mid_value["总务扣款"] = ['', out_qi_ta_statues]

        # 判断AD和现金AD状态

        if AD_statues == "仅有一笔AD":

            if xian_jin_AD_statues == '现金AD未上传':
                AD_result = True
                AD_value = AD[0][2]

            else:
                AD_result = False
        elif AD_statues == "AD未上传":
            if xian_jin_AD_statues == '仅有一笔现金AD':
                AD_result = True
                AD_value = xian_jin_AD[0][2]
            else:
                AD_result = False

        else:
            AD_result = False
            AD_value = 0

        if AD_result and zhong_jie_fei_statues == "仅有一笔中介费" and fan_jin_statues == "仅有一笔返金":
            try:
                if (int(zhongjiefei[0][2]) + int(AD_value) + 770) * k >= int(fan_jin[0][2]):
                    mid_value["对比结果"] = "返金符合要求"
                else:
                    mid_value["对比结果"] = "返金不符合要求"
            except:
                mid_value["对比结果"] = "可能存在多笔AD或同时存在AD和现金AD"
        elif fan_jin_statues == "此单无返金":
            mid_value["对比结果"] = "此单无返金"
        elif zhong_jie_fei_statues == "中介费未上传" or "无中介费":
            try:
                fan_jin = fan_jin[0][2]
            except:
                fan_jin = 0
            try:
                if (int(AD_value) + 770) * k >= int(fan_jin):
                    mid_value["对比结果"] = "返金符合要求"
                else:
                    mid_value["对比结果"] = "返金不符合要求"
            except:
                mid_value["对比结果"] = "可能存在多笔AD或同时存在AD和现金AD"
        else:
            mid_value["对比结果"] = "缺少输入项"

        return mid_value

    def check_angency_fee(self, driver, date):

        if self.cai_wu_statues["中介费"][1] == "仅有一笔中介费":
            for i in range(len(self.values)):
                if self.values[i][0] == "中介费":
                    id = self.values[i][7]
                    check_person = self.values[i][6]
                    driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/" + str(self.money_id) + ".html")
                    btns = driver.find_elements_by_tag_name("button")
            if check_person == "":
                for i in range(len(btns)):
                    if btns[i].get_attribute("data-target") == "#exampleModal" + str(id):
                        btns[i].click()
                        time.sleep(1)
                        # 点击审核成功
                        driver.find_element_by_xpath(
                            "/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal" + id + "']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][1]/label[@class='radio-inline'][1]/input").click()
                        # 输入日期
                        date_btn = driver.find_element_by_xpath(
                            "/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal" + id + "']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][3]/div[@class='input-group date form_date col-md-12']/input[@class='form-control']")
                        date_btn.click()
                        date_btn.clear()
                        date_btn.send_keys(date)
                        date_btn.send_keys(Keys.ENTER)

                        # 点击提交
                        driver.find_element_by_xpath(
                            "/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal" + id + "']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-footer']/button[@class='btn btn-primary']").click()
                        time.sleep(5)
                        break

    def check_office_fee(self, driver, date):

        if self.cai_wu_statues["总务扣款"][1] == "仅有一笔总务扣款":

            for i in range(len(self.values)):
                if self.values[i][1] == "出金":
                    if self.values[i][0] == "其他":
                        id = self.values[i][7]
                        check_person = self.values[i][6]

                    driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/" + str(self.money_id) + ".html")
                    btns = driver.find_elements_by_tag_name("button")
            if check_person == "":
                for i in range(len(btns)):

                    if btns[i].get_attribute("data-target") == "#exampleModal" + str(id):
                        btns[i].click()
                        time.sleep(1)
                        # 点击审核成功
                        driver.find_element_by_xpath(
                            "/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal" + id + "']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][1]/label[@class='radio-inline'][1]/input").click()
                        # 输入日期
                        date_btn = driver.find_element_by_xpath(
                            "/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal" + id + "']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][3]/div[@class='input-group date form_date col-md-12']/input[@class='form-control']")
                        date_btn.click()
                        date_btn.clear()
                        date_btn.send_keys(date)
                        date_btn.send_keys(Keys.ENTER)

                        # 点击提交
                        driver.find_element_by_xpath(
                            "/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal" + id + "']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-footer']/button[@class='btn btn-primary']").click()
                        time.sleep(5)
                        break
            else:
                pass

    def check_fan_jin(self, driver, date):
        if self.cai_wu_statues["对比结果"] == "返金符合要求":
            for i in range(len(self.values)):
                if self.values[i][0] == "返金":
                    id = self.values[i][7]
                    check_person = self.values[i][6]
                    driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/" + str(self.money_id) + ".html")
                    btns = driver.find_elements_by_tag_name("button")
            if check_person == "":
                for i in range(len(btns)):
                    if btns[i].get_attribute("data-target") == "#exampleModal" + str(id):
                        btns[i].click()
                        time.sleep(1)
                        # 点击审核成功
                        driver.find_element_by_xpath(
                            "/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal" + id + "']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][1]/label[@class='radio-inline'][1]/input").click()
                        # 输入日期
                        date_btn = driver.find_element_by_xpath(
                            "/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal" + id + "']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-body']/div[@class='form-group'][3]/div[@class='input-group date form_date col-md-12']/input[@class='form-control']")
                        date_btn.click()
                        date_btn.clear()
                        date_btn.send_keys(date)
                        date_btn.send_keys(Keys.ENTER)

                        # 点击提交
                        driver.find_element_by_xpath(
                            "/html/body[@class='modal-open']/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/div[@id='exampleModal" + id + "']/div[@class='modal-dialog']/div[@class='modal-content']/div[@class='modal-footer']/button[@class='btn btn-primary']").click()
                        time.sleep(5)
                        break

    def check_final(self, driver):
        driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/" + str(self.money_id) + ".html")
        kind = driver.find_element_by_id('selectkhzt')
        kind_lei = Select(kind)
        kind_lei.select_by_index(1)

        btn = driver.find_element_by_xpath(
            "/html/body/div[@class='mm']/div[@class='right']/div[@class='pre-scrollable']/if/form[@id='form']/table[@class='table table-bordered']/tbody/tr/td/div[@class='form-group'][2]/div[@class='col-sm-offset-1 col-sm-3']/button[@class='btn btn-default']")
        btn.click()
        time.sleep(5)

    def get_customer_id(self):
        s = self.login_res("0188", "0123456789")
        html = s.get("http://hy.ayqiandu.net/index/shenqing/index.html?user_id=&khlb=&wjfh=" + str(
            self.number) + "&wujianming=&xm=&glgsdh=&glgsm=&shengri=").text
        soup = BeautifulSoup(html, features="lxml")

        this_number = str(self.number).strip()

        tds = soup.find_all(name="td")

        for i in range(len(tds)):

            if tds[i].get_text().strip() == this_number:
                customer_id = tds[i - 3].get_text()
                # 调整下行的find_next_siblings()[number]中number可以得到对应数据 0为科室，1为社员或兼职，2为名字，3为客户姓名，4为客户状态
                wu_jian_ming = tds[i + 1].get_text()
                guan_li_gong_si = tds[i + 2].get_text()
                client_statues = tds[i + 6].get_text()
                return [customer_id, wu_jian_ming, guan_li_gong_si, client_statues]
        return ['该订单未上传', '', '', '该订单未上传或读取问题']

    def check_user_statues(self, driver):
        driver.get("http://hy.ayqiandu.net/index/caiwu/info/id/" + str(self.money_id) + ".html")
        kind = driver.find_element_by_id('selectkhzt')
        kind_lei = Select(kind)
        kind_lei.select_by_index(1)

        btns = driver.find_elements_by_tag_name("button")
        k = []
        for i in range(len(btns)):
            if btns[i].text == "审核":
                k.append(btns[i])
        btn = k[len(k) - 1]
        btn.click()
        time.sleep(4)

    # 想获取更多客户管理系统就在这里更改吧
    def get_client_system_values(self):

        s = self.login_res("0188", "0123456789")
        html = s.get("http://hy.ayqiandu.net/index/shenqing/info/id/" + str(self.customer_id) + ".html").text
        soup = BeautifulSoup(html, features="lxml")

        apply_money_statues = soup.find_all("select")[4].find_all("option")
        for i in range(len(apply_money_statues)):
            if len(apply_money_statues[i].attrs) == 2:
                apply_statues = apply_money_statues[i].get_text()
        a = soup.find_all("input")

        fan_jin = soup.find_all("input")[27].attrs['value']
        money = soup.find_all("input")[20].attrs['value']
        ps = soup.find_all("input")[22].attrs['value']
        AD_statue = soup.find_all("input")[21].attrs['value']
        angency_fee = soup.find_all("input")[23].attrs['value']
        rent = soup.find_all("input")[34].attrs['value']
        management_fee = soup.find_all("input")[35].attrs['value']
        # 定金状态 定金金额 备注 返金金额 AD金额 中介费 房租 管理费
        return [apply_statues, money, ps, fan_jin, AD_statue, angency_fee, rent, management_fee]

def login_res(name, password):
    login_url = 'http://hy.ayqiandu.net/index/login/index.html'
    formData = {'uname': name, 'passwd': password}
    hea = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) PhantomJS/41.0.2272.118 Safari/537.36'}
    s = requests.Session()
    s.post(login_url, data=formData, headers=hea)
    return s

s=login_res("0188",'0123456789')
result={}
for i in range(1,71):
    html=s.get("http://hy.ayqiandu.net/index/caiwu/index.html?page="+str(i)).text
    soup = BeautifulSoup(html, features="lxml")
    tds=soup.find_all(name="td")
    for j in range(len(tds)):
        if j%11==0:
            number=tds[j+1].get_text()
            if "MAI" in number:
                pass
            elif "レ" in number:
                pass
            else:
                print("第%d页" % i)
                try:
                    company_this=Company(number)
                    result[number]={'中介费':company_this.cai_wu_statues["中介费"][1],'现金AD':company_this.cai_wu_statues["现金AD"][1],'AD':company_this.cai_wu_statues["AD"][1], '返金':company_this.cai_wu_statues["返金"][1], '总务扣款':company_this.cai_wu_statues["总务扣款"][1],"状态":company_this.statute}
                except:
                    pass




def put_cai_wu_value(path,value_content):
    # 打开excel
    name=os.path.join(path)
    wb=excel.load_workbook(filename=name)
    # 获取并打开工作册
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])
    # 创建数据存储
    content_mid=[]
    # 遍历行获取数据并输出

    # 此处把excel第一行的标题插入
    ws.cell(row=1, column=1, value="单号")
    ws.cell(row=1, column=2, value="中介费")
    ws.cell(row=1, column=3, value="现金AD")
    ws.cell(row=1, column=4, value="AD")
    ws.cell(row=1, column=5, value="返金")
    ws.cell(row=1, column=6, value="总务扣款")
    ws.cell(row=1, column=6, value="状态")


    keys=list(value_content.keys())
    print(keys)
    for i in range(len(keys)):


        # 开始循环写入数据

        ws.cell(row = i+2, column = 1,value=keys[i])
        ws.cell(row = i+2, column = 2,value=value_content[keys[i]]["中介费"])
        ws.cell(row = i+2, column = 3,value=value_content[keys[i]]["现金AD"])
        ws.cell(row=i + 2, column=4, value=value_content[keys[i]]["AD"])
        ws.cell(row=i + 2, column=5, value=value_content[keys[i]]["返金"])
        ws.cell(row=i + 2, column=6, value=value_content[keys[i]]["总务扣款"])
        ws.cell(row=i + 2, column=7, value=value_content[keys[i]]["状态"])




        #value_content.append(k)
    wb.save(path)



put_cai_wu_value("C:\\Users\\user\\Desktop\\单子详情.xlsx",result)
