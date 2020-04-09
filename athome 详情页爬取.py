# -*- coding:utf-8 -*-

import socket,re,os,random
import socks
import requests
from bs4 import BeautifulSoup
from stem import Signal
from stem.control import Controller
import threading
import queue as Queue
import openpyxl as excel

controller = Controller.from_port(port=9151)
controller.authenticate()

socks.set_default_proxy(socks.SOCKS5, "127.0.0.1", 9150)
socket.socket = socks.socksocket
# urls用于存放所有网页地址
urls=[]

result=[]
def get_excel_value(path,value_list):
    # 打开excel
    name=os.path.join(path)
    wb=excel.load_workbook(filename=name)
    # 获取并打开工作册
    sheets = wb.get_sheet_names()
    print(sheets)
    ws = wb.get_sheet_by_name(sheets[0])
    # 创建数据存储
    content_mid=[]
    # 遍历行获取数据并输出
    rows = ws.rows
    for row in rows:
        k=[] # 注意这里 放在第二个循环内是不可以的

        for cell in row:
            k.append(cell.value)


        value_list.append(k)

get_excel_value("C:\\Users\\user\\Desktop\\网址.xlsx",urls)


user_agent = [
"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.2; WOW64) AppleWebKit/537.36 (KHTML like Gecko) Chrome/44.0.2403.155 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36",
"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.1 Safari/537.36",
"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.0 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2227.0 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2226.0 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.4; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2225.0 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.3; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2225.0 Safari/537.36",
"Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2224.3 Safari/537.36",
"Mozilla/5.0 (Windows NT 10.0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/40.0.2214.93 Safari/537.36",
"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/37.0.2062.124 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/37.0.2049.0 Safari/537.36",
"Mozilla/5.0 (Windows NT 4.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/37.0.2049.0 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.67 Safari/537.36",
"Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.67 Safari/537.36",
"Mozilla/5.0 (X11; OpenBSD i386) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1985.125 Safari/537.36",
"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/36.0.1944.0 Safari/537.36",
"Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.3319.102 Safari/537.36",
"Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.2309.372 Safari/537.36",
"Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.2117.157 Safari/537.36",
"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_3) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/35.0.1916.47 Safari/537.36",
"Mozilla/5.0 (Windows NT 5.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1866.237 Safari/537.36",
"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.137 Safari/4E423F",
"Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/33.0.1750.517 Safari/537.36",
"Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/32.0.1667.0 Safari/537.36",
]


# 头设置
headers = {
    'Accept': '*/*',
    'Accept-Language': 'en-US,en;q=0.8',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',

}
def get_values(url):
    link=url
    controller.signal(Signal.NEWNYM)
    this_header=headers
    this_header['User-Agent']=user_agent[random.randint(0,len(user_agent)-1)]
    a = requests.get(link, headers=this_header,timeout=20)
    html = a.text.encode(a.encoding).decode("utf-8")
    soup=BeautifulSoup(html, features="lxml")
    meta = soup.find_all("meta")
    if "name" in meta[0].attrs:
        if meta[0]["name"] == "ROBOTS":
            get_values(url)
    else:
        try:
            jiao_tong=soup.find(name="th",text="交通").find_next(name="td").get_text().replace("\n","").replace("\t","").replace("（電車ルート案内）","")
            if soup.find(name="th",text="その他交通"):
                others_jiao_tong=soup.find(name="th",text="その他交通").find_next(name="td").get_text().replace("\n","").replace("\t","").replace("（電車ルート案内）","")
            else:
                others_jiao_tong =""
            address=soup.find(name="th",text="所在地").find_next(name="td").get_text().replace("の賃料 家賃相場","").replace("\n","").replace("\t","")
            kind=soup.find(name="th",text="物件種目").find_next(name="td").get_text().replace("\n","").replace("\t","")
            rent=soup.find(name="th",text="賃料").find_next(name="td").get_text().replace("\n","").replace("\t","")
            guan_li_fei = soup.find(name="th", text="管理費等").find_next(name="td").get_text().replace("\n","").replace("\t","")
            fu_jin_and_bao_zheng_jin=soup.find(name="th", text="敷金/保証金").find_next(name="td").get_text().replace("\n","").replace("\t","")
            li_jin=soup.find(name="th", text="礼金").find_next(name="td").get_text().replace("\n","").replace("\t","")
            wu_jin_ming=soup.find(name="th", text="建物名・部屋番号").find_next(name="td").get_text().replace("\n","").replace("\t","")
            fang_xing=soup.find(name="th", text="間取り").find_next(name="td").get_text().replace("\n","").replace("\t","")
            mian_jin=soup.find(name="th", text="専有面積").find_next(name="td").get_text().replace("\n","").replace("\t","")
            lou_ceng = soup.find(name="th", text="階建 / 階").find_next(name="td").get_text().replace("\n","").replace("\t","")
            jie_gou=soup.find(name="th", text="建物構造").find_next(name="td").get_text().replace("\n","").replace("\t","")
            jian_zao_ri=soup.find(name="th", text="築年月").find_next(name="td").get_text().replace("\n","").replace("\t","")
            zuo_biao=soup.find(id="MAP").attrs
            lat=zuo_biao["lat"]
            lon=zuo_biao["lon"]
            out=[jiao_tong,others_jiao_tong,address,kind,rent,guan_li_fei,fu_jin_and_bao_zheng_jin,li_jin,wu_jin_ming,fang_xing,
                 mian_jin,lou_ceng,jie_gou,jian_zao_ri,lat,lon]

            return out
        except:

            return None

def put_cai_wu_value(path,value_list):
    # 打开excel
    name=os.path.join(path)
    wb=excel.load_workbook(filename=name)
    # 获取并打开工作册
    sheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheets[0])
    row = ws.max_row
    if value_list is not None:
        for i in range(len(value_list)):
            ws.cell(row=row + 1, column=i + 1, value=value_list[i])
    wb.save(path)

'''
class myThread(threading.Thread):
    def __init__(self,name,q,i):
        threading.Thread.__init__(self)
        self.name=name
        self.q=q
        self.i=i
    def run(self):
        while True:
                #print("开始"+self.name)
                aim=get_values(self.q.get(1))
                if aim is not None:
                    print(aim)
                    result.append(aim)

workQueue=Queue.Queue()
for url in urls[:1000]:
    workQueue.put(url[0])
threads=[]

for i in range(100):
    thread=myThread("线程"+str(i),workQueue,i)
    threads.append(thread)
for t in threads:
    t.start()

for t in threads:
    t.join()



print(result)


'''
k=0
for i in urls:
    url =i[0]
    try:
        aim = get_values(url)

        if aim is not None:
            print(aim)
            k+=1
            print(k)
            put_cai_wu_value("C:\\Users\\user\\Desktop\\athome值.xlsx", aim)
    except:
        pass

print("执行完毕")
